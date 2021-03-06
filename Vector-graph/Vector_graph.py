###Los graficos del script son calculados y graficados en el sentido horario de las manecillas del reloj, ademas de esto se toma la direccion geografica norte
### como punto de inicio de 0 grados, todo esto se realiza de modo que la orientacion de los graficos sea la misma a la orientacion que se usa en la cartografia

#                     N 0°
#                      ^
#                      |
#                      |
#       O 270° <------ + ------>  E 90°
#                      |
#                      | 
#                      v            
#                     S 180°

import os                                                                                                               ## Importar modulo de comandos del sistema windows CMD
import math                                                                                                             ## Importar modulo matematico de python
import numpy as np                                                                                                      ## Libreria con modulos matematicos avanzados
import matplotlib.pyplot as plt                                                                                         ## Libreria para graficacion de datos 
from openpyxl import load_workbook                                                                                      ##  Importa modulo de lectura de archivos (load_workbook ) -->>
from openpyxl import Workbook                                                                                           ##  Importa modulo de lectura de archivos (load_workbook ) -->>                                                                                                                        ##  -->>de la libreria openpyxl


###VARIABLES DE INDENTIFICACION###
estacion = "GARA"
##PPP
Velocidad_E = 1.165
velocidad_N = 13.36
path = "D:/GNSS Project Files/MODEL MOTION PLATE EXCEL/ESTACIONES CA(NNR)/MAGNAECO"
Relacion = "CA(NNR)"
path_OUT = "C:/Users/julia/Desktop/vectores/CA(NNR)/"

residual_Evel = []
residual_Nvel = []
residual_magntd = []
residual_AZM = []
 
#Funciones Matematicas#
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
def magnitude(x1,y1,x2,y2):                                                                                             # X = Coordebadas de longitud Y = Coordenadas de latitud 
    raiz = math.sqrt((x2-x1)**2+(y2-y1)**2)                                                                  #calcula la magnitud de un vector y luego la redondea a dos decimales  
    return raiz                                                                                                       ##devuelve el valor resultante de la operancion
   
def azimuth(y, x):

    rads = math.atan2(x, y)                                                                                             #Calcula la magnitud de un vector, se invierten xy, ya que los angulos->       
    angulo = math.degrees(rads)                                                                               #-> Se grafican en la cartografia, se indican en el sentido horario 
    if angulo < 0:
        angulo = angulo + 360 
        return angulo                         
    else:     
        return angulo                                                                                                 ##devuelve el valor resultante de la operancion
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#

#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
#recorrido angular de los circulos de nivel
an = np.linspace(0, 2 * np.pi, 100)
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#

#Funcion de configuracion de subgraficos y sus ejes
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
#CONFIGURACION DE ATRIBUJTOS DE EJES DE LOS GRAFICOS
def ejes_conf(N_grafico,title):
    s = plt.subplot(1, 3, N_grafico)                                                                                  #Creacion de subgraficos en formato objetomediante la funcion subplot
    #Texto de los puntos cardinales
    direccion = ['N','S','E','O']
    plt.xlim(-21.3, 21.3)                                                                                             #Estabkece los limites del eje X
    plt.ylim(-21.3, 21.3)                                                                                             #Estabkece los limites del eje Y
    #Configuracion del ubcacion de los ejes XY                                                                                                                                                                                                                                 #Crea el objeto ax de la clase axes con las caracteristicas del eje
    s.spines['right'].set_color('none')                                                                               #Le quita el color a la linea del borde derecho del marco del eje
    s.spines['top'].set_color('none')                                                                                 #Le quita el color a la linea del borde superior del marco del eje
    s.xaxis.set_ticks_position('bottom')                                                                              #Selecciona el eje inferior x para que su posiciohn sea ajustada
    s.spines['bottom'].set_position(('data',0))                                                                       # Se Ajusta la interseccion del eje x al punto 0
    s.yaxis.set_ticks_position('left')                                                                                #Selecciona el eje inferior y para que su posicion sea ajustada                                                                                 #Selecciona el eje inferior x para ser ajustado
    s.spines['left'].set_position(('data',0))                                                                         # Se Ajusta la interseccion del eje y al punto 0
    ##Graficar circulos de nivel
    plt.plot(21.3 * np.cos(an), 21.3 * np.sin(an), color='#BBBBBB', linestyle="-.")                                   #Graficacion de los circulos de nivel
    for i in range(0,21,5): plt.plot(i * np.cos(an), i * np.sin(an), color='#BBBBBB', linestyle="-.")                 #Las lines angulares se calcularon mediante el uso del modulo numpy
    #Escritura de los Puntos cardinales 
    plt.text(1,20.3, direccion[0], family='serif', style='italic', ha='right', wrap=True, size=15)                    
    plt.text(1,-21, direccion[1], family='serif', style='italic', ha='right', wrap=True, size=15) 
    plt.text(21, 0.3, direccion[2], family='serif', style='italic', ha='right', wrap=True, size=15) 
    plt.text(-20.3, 0.3, direccion[3], family='serif', style='italic', ha='right', wrap=True, size=15)      
    s.set_title(title)
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
    
### Funcion para graficacion de vetores de los modelos de movmiento ### Solo se puede usar despues de cargar los archivos .xlsx 
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
def Vmodel_graph(MotionModel,PPP_E, PPP_N):
    counter = 0
    global azimuths_global                                                                                                 #lista de tipo global para ser instanciada fuera de la funcion
    azimuths_global = []                                                                                                   #Asignacion de valores  de azimuth a la variable global
    global magnt_global                                                                                                    #lista de tipo global para ser instanciada fuera de la funcion
    magnt_global = []

    for model in MotionModel:
        print(model[4])
        Evel = sheet_ranges[model[2]].value                                                                                # Obtencion de los datos  de Evel de las celdas dentro de excel
        Nvel = sheet_ranges[model[3]].value                                                                                # Obtencion de los datos  de Nvel de las celdas dentro de excel
        
        #Calculo de magnitud y azimuth 
        print("Evel: " + str(Evel) + " " + "Nvel: " + str(Nvel))
        magnitud = magnitude(0, 0, Evel, Nvel)
        grados = azimuth(Nvel, Evel)

        azimuths_global.append(round(grados,2))
        magnt_global.append(round(magnitud,2))

        #Graficacion del vector 
        plt.arrow(0, 0, Evel, Nvel, head_width=0.40, head_length=0.80, width = 0.15,  fc=model[5], ec='#000000')    

        #Graficacion de la leyenda del vector
        plt.arrow(16.2, 20.25-counter, 1, 0, head_width=0.30, head_length=0.60, width = 0.10,  fc=model[5], ec='#000000')  #flecha de leyenda 
        plt.text(22.5,20-counter, model[4], family='serif', style='italic', ha='center', wrap=True, size=10)               #Texto de leyenda
                                                                                       
        counter = counter + 1
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
                   
### CALCULO Y GRAFICACION VECTOR INDIVIDUAL PPP ###
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#           
def PPP_Vector(long, latd):
    print("PPP")
    ##Calculo de magnitud y azimuth del vector PPP                                        
    magnitud = magnitude(0, 0, long, latd)
    grados = azimuth(latd, long)

    #Graficacion del vector                                                                                               #Se le asignan los atributos de coordenadas y stilo al objeto flecha.
    plt.arrow(0, 0, long, latd, head_width=0.4, head_length=0.8, width = 0.15,  fc='#000000', ec='#000000')                   #(X, ,Y, unidades a recorrer en x,unidades a recorrer en Y)  
                                                                                                                          #fc = color relleno de la flecha ec = color del borde de la flecha.   
    #LEYENDA PPP
    plt.arrow(16.2, 21.1, 1, 0, head_width=0.30, head_length=0.60, width = 0.10,  fc='#000000', ec='#000000')                    #flecha de leyenda 
    plt.text(22.5, 21, "PPP", family='serif', style='italic', ha='center', wrap=True, size=10)                              #Texto de leyenda
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
                              
##Valores indicativos de las celdas de los modelos de movmiento de placas tectonicas

#Modelos de tipo geodesico
                    #CELDAS DE LECTURA EXCEL UNAVCO                    #CELDAS DE ESCRITURA EXCEL PROCESADO
##LISTA           Long    Lat  Evel Nvel   Modelo de movimiento Color  AZM   MAGNITUD POSICION MODELO
ITRF2008       = ['B12','C12','D11','E11',      "ITRF2008",  '#8F0724'] #Modelo Geodesico       
ITRF2000AS     = ['B24','C24','D23','E23',    "ITRF2000AS",  '#FF0000'] #Modelo Geodesico
ITRF2000DA     = ['B30','C30','D29','E29',    "ITRF2000DA",  '#FF8000'] #Modelo Geodesico
APKIM2005_DGFI = ['B14','C14','D13','E13',"APKIM2005_DGFI",  '#86D200'] #Modelo Geodesico
APKIM2005_IGN  = ['B16','C16','D15','E15', "APKIM2005_IGN",  '#29D200'] #Modelo Geodesico
APKIM2000      = ['B28','C28','D27','E27',     "APKIM2000",  '#008C2B'] #Modelo Geodesico
CGPS2004       = ['B20','C20','D19','E19',      "CGPS2004",  '#00FFD4'] #Modelo Geodesico
REVEL2000      = ['B22','C22','D21','E21',     "REVEL2000",  '#00A7E5'] #Modelo Geodesico
GEODVEL2010    = ['B8','C8','D7','E7',       "GEODVEL2010",  '#0061E5'] #Modelo Geodesico

#Modelo Matricial para seleccion de modelo de movmiento de placa  y seleccion de sus celdas correspodientes en excel
MotionModel_Geodesic = [ITRF2008,ITRF2000AS,ITRF2000DA,APKIM2005_DGFI,APKIM2005_IGN,APKIM2000,CGPS2004,REVEL2000,GEODVEL2010]

#Modelos de tipo Geofisico
                    #CELDAS DE LECTURA EXCEL UNAVCO                    #CELDAS DE ESCRITURA EXCEL PROCESADO
##LISTA           Long  Lat  Evel Nvel  Modelo de movimiento Color  AZM   MAGNITUD POSICION MODELO
NNR_MORVEL     = ['B6','C6','D5','E5',        "NNR_MORVEL",   "#1AD3C2"] #Modelo Geofisico
HS3_NUVEL1A    = ['B26','C26','D25','E25',   "HS3_NUVEL1A",   "#40008D"] #Modelo Geofisico
HS2_NUVEL1A    = ['B32','C32','D31','E31',   "HS2_NUVEL1A",   "#20008D"] #Modelo Geofisico
NUVEL1A        = ['B34','C34','D33','E33',       "NUVEL1A",   "#C50C66"] #Modelo Geofisico
NUVEL1         = ['B36','C36','D35','E35',        "NUVEL1",   "#FC1EFF"] #Modelo Geofisico

#Modelo Matricial para seleccion de modelo de movmiento de placa  y seleccion de sus celdas correspodientes en excel
MotionModel_Geodephysic = [NNR_MORVEL,HS3_NUVEL1A,HS2_NUVEL1A, NUVEL1A, NUVEL1]

#Modelos de movimiento de tipo combinado
                    #CELDAS DE LECTURA EXCEL UNAVCO                    #CELDAS DE ESCRITURA EXCEL PROCESADO
##LISTA            Long  Lat Evel Nvel  Modelo de movimiento Color  AZM   MAGNITUD POSICION MODELO
GSMR2_1        = ['B4','C4','D3','E3',           "GSMR2_1", "#565656"] #Modelo Combinado
GSMR1_2        = ['B18','C18','D17','E17',       "GSMR1_2", "#999999"] #Modelo Combinado
MORVEL2010     = ['B10','C10','D9','E9',      "MORVEL2010", "#7EBBA0"] #Modelo Combinado

#Modelo Matricial para seleccion de modelo de movmiento de placa  y seleccion de sus celdas correspodientes en excel
MotionModel_Combinated =[GSMR2_1,GSMR1_2,MORVEL2010]

### UBUCACION DE DIRECTORIO DE TRABAJO DEL PROGRAMA
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
os.chdir(path)                             ##Asignacion de la ruta donde se encuentran los archivos xlsx
ruta = os.getcwd()                                                                                                 #Obtiene de ubicacion de ruta actual en el programa
print ("DIRECCION: " + ruta)                                                                                       ##imprime la direccion actual del programa mediante el metodo getcwd()
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#

wb = load_workbook(estacion+'_' + Relacion + '.xlsx')                                                                       ##Comando para cargar  el archivo.xlsx
sheetname = str(wb.get_sheet_names())                                                                              ##Comando para obtener el nombre de las hojas de calculo del archivo y convertirlo en string
sheetname = sheetname[2:-2]                                                                                        ##Se ajusta el nombre del sheetname ya que viene con estos caracteres de mas ['sheetname']
print ('El archuvo ' + sheetname +'.xlsx ha sido abierto')  
sheet_ranges = wb[sheetname]                                                                                       ##De esta forma se guarda la hoja de calculo como un objeto y sus celdas como atributos del objeto
print(sheet_ranges)

##Configuracion de la ventana y tipo de grafico a usar
plt.figure('VECTORES_'+estacion + "_" +Relacion,figsize=(30, 8), dpi=80)                                           #Configura el tamaño y la resolucion de la ventana del grafico
###Creacion de fichero que contendra los datos resultantes al procesamiento
fichero = open(path_OUT + "VECTORES_" + estacion + "_" +  Relacion + ".txt", "w")                                  #se crea un archivo fichero .txt en la direccion indicada 

vector_book = Workbook()                                                                                           # Crea el objeto libro de la clase workbook de openpyxl
Vector_sheet = vector_book.create_sheet("Mysheet", 0)                                                              # crea el objeto hoja de la clase hoja 
Vector_sheet.title = estacion                                                                                      # Asigna un string al atributo titulo del objeto hoja

###Sub grafico 1
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
#Creacion de subgrafico en una figura y configuracion de sus ejes
ejes_conf(1, ("Modelos Geodesicos - Vectice: " + estacion))           
PPP_Vector(Velocidad_E,velocidad_N)                                                                                #Funcion de graficacion y calculo del vector PPP            
Vmodel_graph(MotionModel_Geodesic, Velocidad_E,velocidad_N)                                                        #Funcion de graficacion y calculo de los vectores de los modelos de movimiento
#datos ppp en txt
fichero.write("DATOS PROCESADOS PPP \n")
fichero.write("AZM      " + "MAGTD   " + "\n")
fichero.write(str(round(azimuth(velocidad_N, Velocidad_E),2)) + "   " + str(round(magnitude(0, 0, Velocidad_E, velocidad_N),2)) + "\n")
fichero.write("\n")
###Marcadores de columnas archivo txt
fichero.write("MODELOS GEODESICOS - VERTICE: " + estacion + "\n")
fichero.write("AZM     " + "MAGNTD  " + "R-Evel " +	"R-Nvel " + "R-AZM   " + "R-MANGTD  " + "MODELO" + "\n")
###Marcadores de columnas archivo excel
Vector_sheet["A1"] = "DATOS";  Vector_sheet["B1"] = "PPP" 
Vector_sheet["A2"] = "AZM";  Vector_sheet["B2"] = "MAGNTD" 
Vector_sheet["A3"] = round(azimuth(velocidad_N, Velocidad_E),2); Vector_sheet["B3"] = round(magnitude(0, 0, Velocidad_E, velocidad_N),2) ##Inserta los valores del vector PPP en las celdas del archivo xlsx

Vector_sheet["A5"] = "MODELOS"; Vector_sheet["B5"] = "GEODESICOS"
Vector_sheet["A6"] = "AZIMUTH";  Vector_sheet["B6"] = "MANGTD" ;  
Vector_sheet["C6"] = "R-Evel"; Vector_sheet["D6"] = "R-Nvel" ; 
Vector_sheet["E6"] = "R-AZM" ; Vector_sheet["F6"] = "R-MANGTD";
Vector_sheet["G6"] = "MODELO"

counterlist = 0
for listModel in MotionModel_Geodesic:
    ##Calculo de vector residual
    residual_Evel.append(Velocidad_E -  sheet_ranges[listModel[2]].value)                                      # Obtencion de los datos  de velocidad_E del PPP de las celdas dentro de excel 
    residual_Nvel.append(velocidad_N -  sheet_ranges[listModel[3]].value)                                      # Obtencion de los datos  de velocidad_N del PPP de las celdas dentro de excel
    residual_magntd.append(round(magnitude(0, 0, residual_Evel[counterlist],residual_Nvel[counterlist]),2))
    residual_AZM.append(round(azimuth(residual_Nvel[counterlist], residual_Evel[counterlist]),2))              ##Se calcula la del vector residual azimuth en el sentido de las manecillas de reloj
    if residual_magntd[counterlist] == round(magnitude(0, 0, Velocidad_E, velocidad_N),2):                     ##Si el vector del modelo tiene diferencia de magnitud igual a la del PPP
        residual_AZM[counterlist] = (azimuth(velocidad_N, Velocidad_E))                                        ##El angulo se ajustara para que tenga el mismo azimuth que el PPP
                                                                                                               ##Debido a que se genera un error en el azimuth cuando el modelo no tiene magnitud
        
    #Escritura DE DATOS en excel
    Vector_sheet.cell(row = counterlist+7, column = 1).value = azimuths_global[counterlist]
    Vector_sheet.cell(row=counterlist+7, column=2).value = magnt_global[counterlist]
    Vector_sheet.cell(row=counterlist+7, column=3).value = residual_Evel[counterlist]
    Vector_sheet.cell(row=counterlist+7, column=4).value = residual_Nvel[counterlist] 
    Vector_sheet.cell(row=counterlist+7, column=5).value = residual_AZM[counterlist] 
    Vector_sheet.cell(row=counterlist+7, column=6).value = residual_magntd[counterlist]
    Vector_sheet.cell(row=counterlist+7, column=7).value = listModel[4]
    fichero.write(str(azimuths_global[counterlist]) + "   " + str(magnt_global[counterlist]) + "   "  + str(round(residual_Evel[counterlist],2)) + "   "   
    + str(round(residual_Nvel[counterlist],2)) + "   "  + str(round(residual_AZM[counterlist],2)) + "   "  + str(residual_magntd[counterlist]) + "   "  + listModel[4] + "\n")       #Escritura en .txt
    counterlist = counterlist+1
fichero.write("\n")
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#

###Sub grafico 2
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
#Creacion de subgrafico en una figura y configuracion de sus ejes
ejes_conf(2, ("Modelos Geofisicos - Vectice: " + estacion))
PPP_Vector(Velocidad_E,velocidad_N)                     
Vmodel_graph(MotionModel_Geodephysic, Velocidad_E,velocidad_N)
###Marcadores de columnas archivo TXT
fichero.write("MODELOS GEOFISICOS VERTICE: " + estacion + "\n")
fichero.write("AZM     " + "MAGNTD  " + "R-Evel " +	"R-Nvel " + "R-AZM   " + "R-MANGTD  " + "MODELO" + "\n")
###Marcadores de columnas archivo excel
Vector_sheet["A17"] = "MODELOS"; Vector_sheet["B17"] = "GEOFISICOS"
Vector_sheet["A18"] = "AZIMUTH";  Vector_sheet["B18"] = "MANGTD" ;  
Vector_sheet["C18"] = "R-Evel"; Vector_sheet["D18"] = "R-Nvel" ; 
Vector_sheet["E18"] = "R-AZM" ; Vector_sheet["F18"] = "R-MANGTD";
Vector_sheet["G18"] = "MODELO"

##Reinocio de variables de ciclo
counterlist = 0
residual_Evel = []
residual_Nvel = []
residual_magntd = []
residual_AZM = []
for listModel in MotionModel_Geodephysic:
    ##Calculo de vector residual
    residual_Evel.append(Velocidad_E -  sheet_ranges[listModel[2]].value)                                      # Obtencion de los datos  de velocidad_E del PPP de las celdas dentro de excel 
    residual_Nvel.append(velocidad_N -  sheet_ranges[listModel[3]].value)                                      # Obtencion de los datos  de velocidad_N del PPP de las celdas dentro de excel
    residual_magntd.append(round(magnitude(0, 0, residual_Evel[counterlist],residual_Nvel[counterlist]),2))    ##Se calcula la del vector residual azimuth en el sentido de las manecillas de reloj
    residual_AZM.append(round(azimuth(residual_Nvel[counterlist], residual_Evel[counterlist]),2))
    if residual_magntd[counterlist] == round(magnitude(0, 0, Velocidad_E, velocidad_N),2):                     ##Si el vector del modelo tiene diferencia de magnitud igual a la del PPP
        residual_AZM[counterlist] = (azimuth(velocidad_N, Velocidad_E)-180)                                    ##El angulo se ajustara para que tenga el mismo azimuth que el PPP
                                                                                                               ##Debido a que se genera un error en el azimuth cuando el modelo no tiene magnitud
                                                                                                               ##Este proceso hara que el vector residual sea igual al PPP en magnitud y azimuth
    #Escritura DE DATOS en excel
    Vector_sheet.cell(row = counterlist+19, column = 1).value = azimuths_global[counterlist]
    Vector_sheet.cell(row=counterlist+19, column=2).value = magnt_global[counterlist]
    Vector_sheet.cell(row=counterlist+19, column=3).value = residual_Evel[counterlist]
    Vector_sheet.cell(row=counterlist+19, column=4).value = residual_Nvel[counterlist] 
    Vector_sheet.cell(row=counterlist+19, column=5).value = residual_AZM[counterlist]
    Vector_sheet.cell(row=counterlist+19, column=6).value = residual_magntd[counterlist]
    Vector_sheet.cell(row=counterlist+19, column=7).value = listModel[4]
    fichero.write(str(azimuths_global[counterlist]) + "   " + str(magnt_global[counterlist]) + "   "  + str(round(residual_Evel[counterlist],2)) + "   "   
    + str(round(residual_Nvel[counterlist],2)) + "   "  + str(round(residual_AZM[counterlist],2)) + "   "  + str(residual_magntd[counterlist]) + "   "  + listModel[4] + "\n")       #Escritura en .txt
    counterlist = counterlist+1
fichero.write("\n")

#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
###Sub grafico 3
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
#Creacion de subgrafico en una figura y configuracion de sus ejes
ejes_conf(3, ("Modelos Combinados - Vectice: " + estacion))
PPP_Vector(Velocidad_E,velocidad_N)                     
Vmodel_graph(MotionModel_Combinated, Velocidad_E,velocidad_N)
fichero.write("MODELOS COMBINADOS VERTICE: " + estacion + "\n")
fichero.write("AZM     " + "MAGNTD  " + "R-Evel " +	"R-Nvel " + "R-AZM   " + "R-MANGTD  " + "MODELO" + "\n")
###Marcadores de columnas archivo excel
Vector_sheet["A25"] = "MODELOS"; Vector_sheet["B25"] = "COMBINADO"
Vector_sheet["A26"] = "AZIMUTH";  Vector_sheet["B26"] = "MANGTD" ;  
Vector_sheet["C26"] = "R-Evel"; Vector_sheet["D26"] = "R-Nvel" ; 
Vector_sheet["E26"] = "R-AZM" ; Vector_sheet["F26"] = "R-MANGTD";
Vector_sheet["G26"] = "MODELO"

##Reinocio de variables de ciclo
counterlist = 0
residual_Evel = []
residual_Nvel = []
residual_magntd = []
residual_AZM = []

for listModel in MotionModel_Combinated:
    ##Calculo de vector residual
    residual_Evel.append(Velocidad_E -  sheet_ranges[listModel[2]].value)                                      # Obtencion de los datos  de velocidad_E del PPP de las celdas dentro de excel 
    residual_Nvel.append(velocidad_N -  sheet_ranges[listModel[3]].value)                                      # Obtencion de los datos  de velocidad_N del PPP de las celdas dentro de excel
    residual_magntd.append(round(magnitude(0, 0, residual_Evel[counterlist],residual_Nvel[counterlist]),2))
    residual_AZM.append(round(azimuth(residual_Nvel[counterlist], residual_Evel[counterlist]),2))              ##Se calcula la del vector residual azimuth en el sentido de las manecillas de reloj
    if residual_magntd[counterlist] == round(magnitude(0, 0, Velocidad_E, velocidad_N),2):                     ##Si el vector del modelo tiene diferencia de magnitud igual a la del PPP
        residual_AZM[counterlist] = (azimuth(velocidad_N, Velocidad_E)-180)                                    ##El angulo se ajustara para que tenga el mismo azimuth que el PPP
                                                                                                               ##Debido a que se genera un error en el azimuth cuando el modelo no tiene magnitud
                                                                                                               ##Este proceso hara que el vector residual sea igual al PPP en magnitud y azimuth
    #Escritura DE DATOS en excel
    Vector_sheet.cell(row=counterlist+27, column = 1).value = azimuths_global[counterlist]
    Vector_sheet.cell(row=counterlist+27, column=2).value = magnt_global[counterlist]
    Vector_sheet.cell(row=counterlist+27, column=3).value = residual_Evel[counterlist]
    Vector_sheet.cell(row=counterlist+27, column=4).value = residual_Nvel[counterlist] 
    Vector_sheet.cell(row=counterlist+27, column=5).value = residual_AZM[counterlist] 
    Vector_sheet.cell(row=counterlist+27, column=6).value = residual_magntd[counterlist]
    Vector_sheet.cell(row=counterlist+27, column=7).value = listModel[4]  
    fichero.write(str(azimuths_global[counterlist]) + "   " + str(magnt_global[counterlist]) + "   "  + str(round(residual_Evel[counterlist],2)) + "   "   
    + str(round(residual_Nvel[counterlist],2)) + "   "  + str(round(residual_AZM[counterlist],2)) + "   "  + str(residual_magntd[counterlist]) + "   "  + listModel[4] + "\n")       #Escritura en .txt
    counterlist = counterlist+1
fichero.write("\n")
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#

vector_book.save(path_OUT + "VECTORES_" + estacion + "_" + Relacion + ".xlsx")                                                          # Guarda el archivo xlsx en una direccion especifica
#vector_book.close()
fichero.close() 
plt.grid(False)                                                                                                       #Se omite la graficacion de la grilla
plt.show()                                                                                                            #Se crea muestra el grafico

#plt.savefig(path_OUT + "Vectores_Estacion: " + estacion + ".png", dpi = 80)                                          # Guardar la figura usando 72 puntos por pulgada